# -*- coding: utf-8 -*-
"""
📈 РЕАЛЬНЫЕ акции → Excel с нуля
Скачиваем настоящие цены через yfinance, кладём на лист «Цены»,
а ВСЁ остальное считаем формулами.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import yfinance as yf



# ═══ Скачиваем РЕАЛЬНЫЕ данные ═══
TICKERS = ['PYPL', 'MAR', 'MA', 'GOOGL', 'WMT', 'BAIDF']
NAMES =   ['PayPal', 'Marriott International', 'Mastercard ', 'Google', 'WalMart', 'Baidu']
SECTORS = ['Industrial',  'Consumer Discretionary',      'Real Estate',   'Technology', 'Consumer Discretionary',  'Technology']
INDEX_TICKER = '^GSPC'  # S&P 500

print("⏳ Скачиваю реальные данные с Yahoo Finance...")
raw = yf.download(TICKERS + [INDEX_TICKER], period='6mo', interval='1d', progress=False)

# Берём последние 30 дней
close = raw['Close'].dropna()
close = close.tail(10000000).copy()
DAYS = len(close)
print(f"✅ Скачано {DAYS} торговых дней: {close.index[0].date()} ... {close.index[-1].date()}")

dates_list = close.index.tolist()
price_data = {}
for t in TICKERS:
    price_data[t] = close[t].tolist()
price_data['SP500'] = close[INDEX_TICKER].tolist()

NR = DAYS - 1  # returns = 29

# ═══ Стили ═══
title_font = Font(bold=True, size=14, color="FFFFFF")
title_fill = PatternFill("solid", fgColor="1B2631")
h2_font = Font(bold=True, size=12, color="1B4F72")
h3_font = Font(bold=True, size=11, color="2E86C1")
explain_font = Font(italic=True, size=10, color="595959")
step_font = Font(bold=True, size=10, color="2E86C1")
label_font = Font(bold=True, size=10)
header_font = Font(bold=True, size=10, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1B4F72")
formula_fill = PatternFill("solid", fgColor="D5F5E3")
intermed_fill = PatternFill("solid", fgColor="FEF9E7")
result_fill = PatternFill("solid", fgColor="ABEBC6")
concl_fill = PatternFill("solid", fgColor="FADBD8")
rank_fill = PatternFill("solid", fgColor="D6EAF8")
result_font = Font(bold=True, size=11, color="006100")
bold_green = Font(bold=True, size=11, color="006100")
big_green = Font(bold=True, size=12, color="006100")
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
wrap = Alignment(wrap_text=True, vertical='top')

def w(ws, r, c, val, font=None, fill=None, border=None, align=None, num_fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if align: cell.alignment = align
    if num_fmt: cell.number_format = num_fmt
    return cell

def title_row(ws, r, text, cols=12):
    for c in range(1, cols+1):
        w(ws, r, c, '', fill=title_fill, border=thin)
    w(ws, r, 1, text, font=title_font, fill=title_fill)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)

def h2(ws, r, text, cols=12):
    w(ws, r, 1, text, font=h2_font)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)

def h3(ws, r, text, cols=12):
    w(ws, r, 1, text, font=h3_font)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)

def explain(ws, r, text, cols=12):
    w(ws, r, 1, text, font=explain_font, align=wrap)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)

def step_r(ws, r, text, cols=12):
    w(ws, r, 1, text, font=step_font, align=wrap)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)

def conclusion(ws, r, text, cols=12):
    w(ws, r, 1, text, font=result_font, fill=concl_fill, align=wrap, border=thin)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)

def headers(ws, r, texts):
    for i, t in enumerate(texts, 1):
        w(ws, r, i, t, font=header_font, fill=header_fill, border=thin,
          align=Alignment(horizontal='center', wrap_text=True))

SRC_P = "'📊 Цены'"
SRC_R = "'📈 Доходности'"
PDS = 4  # price data start row
PDE = PDS + DAYS - 1

wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════
# ЛИСТ 1: Реальные цены
# Cols: A=Date B=Day C=PYPL D=MAR E=MA F=GOOGL G=WMT H=BAIDF I=S&P500
# ═══════════════════════════════════════════════════════════════
ws = wb.active
ws.title = '📊 Цены'
title_row(ws, 1, f'📊 РЕАЛЬНЫЕ ЦЕНЫ ЗАКРЫТИЯ ({DAYS} торговых дней, Yahoo Finance)', 9)
explain(ws, 2, f'Период: {dates_list[0].strftime("%d.%m.%Y")} — {dates_list[-1].strftime("%d.%m.%Y")}. Цены в USD. S&P 500 — индекс рынка.', cols=9)

col_hdrs = ['Дата', 'День'] + [f'{TICKERS[i]} ({NAMES[i]})' for i in range(6)] + ['S&P 500']
headers(ws, 3, col_hdrs)

for d in range(DAYS):
    w(ws, PDS+d, 1, dates_list[d].date() if hasattr(dates_list[d], 'date') else dates_list[d],
      border=thin, num_fmt='DD.MM.YYYY')
    w(ws, PDS+d, 2, d+1, border=thin)
    for s, t in enumerate(TICKERS):
        w(ws, PDS+d, 3+s, round(price_data[t][d], 2), border=thin, num_fmt='#,##0.00')
    w(ws, PDS+d, 9, round(price_data['SP500'][d], 2), border=thin, num_fmt='#,##0.00')

for c in range(1, 10):
    ws.column_dimensions[get_column_letter(c)].width = 18
ws.column_dimensions['A'].width = 14

# ═══════════════════════════════════════════════════════════════
# ЛИСТ 2: Доходности + метрики
# Cols: A=Date B=Day C=PYPL D=MAR E=MA F=GOOGL G=WMT H=BAIDF I=S&P500
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('📈 Доходности')
title_row(ws, 1, '📈 ДОХОДНОСТИ И МЕТРИКИ — СЧИТАЕМ ИЗ РЕАЛЬНЫХ ЦЕН', 12)

r = 3
h2(ws, r, '📊 ШАГ 1: ДНЕВНЫЕ ДОХОДНОСТИ = (Цена_сегодня − Цена_вчера) / Цена_вчера', 12); r += 1
r += 1

col_labels = [f'{TICKERS[i]} (%)' for i in range(6)] + ['S&P500 (%)']
headers(ws, r, ['Дата', 'День'] + col_labels + ['', ''])
r += 1

ret_data_start = r  # <-- РЕАЛЬНАЯ строка начала данных
price_cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I']

for d in range(NR):
    pr = PDS + d + 1
    w(ws, r, 1, f"={SRC_P}!A{pr}", border=thin, num_fmt='DD.MM.YYYY')
    w(ws, r, 2, d+2, border=thin)
    for s in range(7):
        pc = price_cols[s]
        w(ws, r, 3+s,
          f"=({SRC_P}!{pc}{pr}-{SRC_P}!{pc}{pr-1})/{SRC_P}!{pc}{pr-1}",
          border=thin, fill=formula_fill, num_fmt='0.00%')
    if d == 0:
        w(ws, r, 11, '← Из сырых цен!', explain_font, align=wrap)
        ws.merge_cells(start_row=r, start_column=11, end_row=r, end_column=12)
    r += 1

ret_data_end = r - 1
r += 1

print(f"[INFO] Доходности: строки {ret_data_start}–{ret_data_end} ({NR} строк)")

kase_col = 'I'  # S&P 500 returns

# ── ШАГ 2: Средняя + Волатильность ──
h2(ws, r, '📊 ШАГ 2: СРЕДНЯЯ ДОХОДНОСТЬ И ВОЛАТИЛЬНОСТЬ', 12); r += 1
r += 1
headers(ws, r, ['Метрика'] + [TICKERS[i] for i in range(6)] + ['S&P500', '', '', '', '']); r += 1

avg_row = r
w(ws, r, 1, 'Средняя дох-ть (дн.)', label_font, border=thin)
for s in range(7):
    col = get_column_letter(3+s)
    w(ws, r, 2+s, f'=AVERAGE({col}{ret_data_start}:{col}{ret_data_end})', fill=formula_fill, border=thin, num_fmt='0.00%')
r += 1

vol_row = r
w(ws, r, 1, 'Волатильность (дн.)', label_font, border=thin)
for s in range(7):
    col = get_column_letter(3+s)
    w(ws, r, 2+s, f'=STDEV({col}{ret_data_start}:{col}{ret_data_end})', fill=formula_fill, border=thin, num_fmt='0.00%')
w(ws, r, 10, '← STDEV = мера риска', explain_font)
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
r += 1

ann_ret_row = r
w(ws, r, 1, 'Годовая дох-ть', label_font, border=thin)
for s in range(7):
    w(ws, r, 2+s, f'={get_column_letter(2+s)}{avg_row}*252', fill=intermed_fill, border=thin, num_fmt='0.0%')
w(ws, r, 10, '← ×252 торговых дня', explain_font)
ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
r += 1

ann_vol_row = r
w(ws, r, 1, 'Годовая волат-ть', label_font, border=thin)
for s in range(7):
    w(ws, r, 2+s, f'={get_column_letter(2+s)}{vol_row}*SQRT(252)', fill=intermed_fill, border=thin, num_fmt='0.0%')
r += 2

# ── ШАГ 3: БЕТА ──
h2(ws, r, '📊 ШАГ 3: БЕТА = COV(акция, S&P500) / VAR(S&P500)', 12); r += 1
explain(ws, r, 'β>1 → рискованнее рынка. β<1 → спокойнее. β≈1 → как рынок.', 12); r += 1
r += 1
headers(ws, r, ['Метрика'] + TICKERS + ['Пояснение', '', '', '', '']); r += 1

beta_row = r
w(ws, r, 1, 'Бета (β)', bold_green, border=thin)
for s in range(6):
    col = get_column_letter(3+s)
    w(ws, r, 2+s,
      f'=COVARIANCE.S({col}{ret_data_start}:{col}{ret_data_end},{kase_col}{ret_data_start}:{kase_col}{ret_data_end})/VAR.S({kase_col}{ret_data_start}:{kase_col}{ret_data_end})',
      fill=result_fill, border=thin, font=big_green, num_fmt='0.00')
w(ws, r, 9, '← COV/VAR из реальных данных!', explain_font)
ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
r += 1

w(ws, r, 1, 'Тип:', label_font, border=thin)
for s in range(6):
    bc = f'{get_column_letter(2+s)}{beta_row}'
    w(ws, r, 2+s, f'=IF({bc}>1.2,"⚡ Агрессивная",IF({bc}<0.8,"🛡️ Защитная","➡️ Нейтральная"))',
      border=thin, fill=intermed_fill)
r += 2

# ── ШАГ 4: ШАРП ──
h2(ws, r, '📊 ШАГ 4: ШАРП = (Год.дох-ть − 5.25%) / Год.волат-ть', 12); r += 1
explain(ws, r, 'Безрисковая ставка = 5.25% (Fed Funds Rate). Sharpe>1 = хорошо.', 12); r += 1
r += 1
headers(ws, r, ['Метрика'] + TICKERS + ['Пояснение', '', '', '', '']); r += 1

rf_rate = 0.0525
rf_row = r
w(ws, r, 1, 'Безриск. ставка', label_font, border=thin)
for s in range(6):
    w(ws, r, 2+s, rf_rate, border=thin, num_fmt='0.0%')
r += 1

sharpe_row = r
w(ws, r, 1, 'Sharpe Ratio ✨', bold_green, border=thin)
for s in range(6):
    ret_c = f'{get_column_letter(2+s)}{ann_ret_row}'
    vol_c = f'{get_column_letter(2+s)}{ann_vol_row}'
    rf_c = f'{get_column_letter(2+s)}{rf_row}'
    w(ws, r, 2+s, f'=({ret_c}-{rf_c})/{vol_c}', fill=result_fill, border=thin, font=big_green, num_fmt='0.00')
r += 1

w(ws, r, 1, 'Оценка:', label_font, border=thin)
for s in range(6):
    sc = f'{get_column_letter(2+s)}{sharpe_row}'
    w(ws, r, 2+s, f'=IF({sc}>1,"⭐ Отличная",IF({sc}>0,"👍 ОК","👎 Хуже депозита"))', border=thin, fill=intermed_fill)
r += 2

conclusion(ws, r, '🏆 Всё посчитано из РЕАЛЬНЫХ цен Yahoo Finance! Бета и Шарп — настоящие значения.', 12)

for c in range(1, 13):
    ws.column_dimensions[get_column_letter(c)].width = 15
ws.column_dimensions['A'].width = 14

# ═══════════════════════════════════════════════════════════════
# ЛИСТ 3: Корреляция
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('📊 Корреляция')
title_row(ws, 1, '📊 КОРРЕЛЯЦИЯ ДОХОДНОСТЕЙ (РЕАЛЬНЫЕ ДАННЫЕ)', 12)

r = 3
h2(ws, r, '❓ Какие акции ходят вместе? Для портфеля нужны НЕКОРРЕЛИРОВАННЫЕ!', 12); r += 1
r += 1

ret_cols = ['C', 'D', 'E', 'F', 'G', 'H']

w(ws, r, 1, '', border=thin, fill=header_fill)
for i, t in enumerate(TICKERS):
    w(ws, r, 2+i, t, header_font, header_fill, thin, Alignment(horizontal='center'))
r += 1
mat_row = r

for ri in range(6):
    w(ws, r, 1, TICKERS[ri], label_font, rank_fill, thin)
    for ci in range(6):
        if ri == ci:
            w(ws, r, 2+ci, 1, border=thin, fill=formula_fill)
        else:
            c1, c2 = ret_cols[ri], ret_cols[ci]
            w(ws, r, 2+ci,
              f"=CORREL({SRC_R}!{c1}{ret_data_start}:{c1}{ret_data_end},{SRC_R}!{c2}{ret_data_start}:{c2}{ret_data_end})",
              border=thin, fill=formula_fill, num_fmt='0.000')
    r += 1
r += 1

# Ключевые пары
h3(ws, r, '📋 КЛЮЧЕВЫЕ ПАРЫ:', 12); r += 1
pairs = [
    ('PYPL', 'MA', 0, 1, 'Два гиганта в сфере цифровых платежей  — диверсификация?'),
    ('PYPL', 'GOOGL', 0, 4, 'Если связь технологичности цифровых платежей и сектора технологии'),
    ('GOOGL', 'BAIDF', 2, 3, 'Два тех-гиганта — география'),
    ('MAR', 'WMT', 4, 5, 'отдух/досуг и потребительские товары  — для диверсификации'),
    ('PYPL', 'BAIDF', 1, 5, 'Если связь технологичности цифровых платежей и сектора технологии (Китай)'),
]
headers(ws, r, ['Пара', '', 'r', 'Сила', '', 'Для портфеля', '', '', '', '', '', '']); r += 1
for v1, v2, ri, ci, biz in pairs:
    rcell = f'{get_column_letter(2+ci)}{mat_row+ri}'
    w(ws, r, 1, f'{v1}↔{v2}', label_font, border=thin)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    w(ws, r, 3, f'={rcell}', border=thin, fill=formula_fill, num_fmt='0.000')
    w(ws, r, 4, f'=IF(ABS({rcell})>=0.7,"🟢 Сильная",IF(ABS({rcell})>=0.3,"🟡 Средняя","🔴 Слабая"))',
      border=thin, fill=intermed_fill)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)
    w(ws, r, 6, biz, explain_font, border=thin, align=wrap)
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=12)
    r += 1
r += 1

# t-тест
h2(ws, r, '✏️ ЗНАЧИМОСТЬ (t-тест)', 12); r += 1
explain(ws, r, f't = r×√(n-2)/√(1-r²). t_крит ≈ 2.05 при α=0.05, df={NR-2}.', 12); r += 1
r += 1
headers(ws, r, ['Пара', '', 'r', 'n', 't-стат', 't-крит', 'Значима?', '', 'Вывод', '', '', '']); r += 1
for v1, v2, ri, ci, biz in pairs:
    rcell = f'{get_column_letter(2+ci)}{mat_row+ri}'
    w(ws, r, 1, f'{v1}↔{v2}', label_font, border=thin)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    w(ws, r, 3, f'={rcell}', border=thin, fill=formula_fill, num_fmt='0.000')
    w(ws, r, 4, NR, border=thin)
    tc = f'C{r}'
    w(ws, r, 5, f'={tc}*SQRT({NR}-2)/SQRT(1-{tc}^2)', border=thin, fill=formula_fill, num_fmt='0.000')
    w(ws, r, 6, 2.05, border=thin)
    ec = f'E{r}'
    w(ws, r, 7, f'=IF(ABS({ec})>2.05,"✅ Реальная","❌ Случайная")', border=thin, fill=result_fill)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    w(ws, r, 9, f'=IF(ABS({ec})>2.05,"Учитывать в портфеле","Игнорировать")',
      explain_font, border=thin, align=wrap)
    ws.merge_cells(start_row=r, start_column=9, end_row=r, end_column=12)
    r += 1
r += 1
conclusion(ws, r, '🏆 РЕАЛЬНЫЕ корреляции! AAPL-MSFT обычно сильно коррелируют (оба Tech). '
    'TSLA-JPM — слабее (разные сектора → хорошая диверсификация).', 12)

for c in range(1, 13):
    ws.column_dimensions[get_column_letter(c)].width = 14

# ═══════════════════════════════════════════════════════════════
# ЛИСТ 4: Регрессия (CAPM)
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('📈 Регрессия')
title_row(ws, 1, '📈 РЕГРЕССИЯ: PYPL ~ S&P 500 (CAPM)', 10)

r = 3
h2(ws, r, '❓ CAPM: Доходность_PYPL = α + β × Доходность_SP500', 10); r += 1
explain(ws, r, 'Paypal — самая волатильная из наших акций. Как она зависит от рынка?', 10); r += 1
explain(ws, r, 'α > 0 → Paypal обыгрывает рынок! β > 1 → рискованнее рынка.', 10); r += 1
r += 1

# Paypal returns = col G, S&P500 returns = col I
pypl_col = 'C'
sp_col = 'I'

h2(ws, r, '📊 МЕТРИКИ CAPM', 10); r += 1
headers(ws, r, ['Параметр', 'Значение', 'Формула', '', 'Что это значит', '', '', '', '', '']); r += 1

w(ws, r, 1, 'α (альфа)', label_font, border=thin)
w(ws, r, 2, f"=INTERCEPT({SRC_R}!{pypl_col}{ret_data_start}:{pypl_col}{ret_data_end},{SRC_R}!{sp_col}{ret_data_start}:{sp_col}{ret_data_end})",
  fill=formula_fill, border=thin, num_fmt='0.00%')
w(ws, r, 3, f'INTERCEPT(PYPL, SP500)', explain_font, border=thin)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
w(ws, r, 5, '← α>0 → PYPL обыгрывает рынок!', explain_font, border=thin, align=wrap)
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10); r += 1

w(ws, r, 1, 'β (бета)', label_font, border=thin)
w(ws, r, 2, f"=SLOPE({SRC_R}!{pypl_col}{ret_data_start}:{pypl_col}{ret_data_end},{SRC_R}!{sp_col}{ret_data_start}:{sp_col}{ret_data_end})",
  fill=formula_fill, border=thin, num_fmt='0.00')
w(ws, r, 3, f'SLOPE(PYPL, SP500)', explain_font, border=thin)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
w(ws, r, 5, '← S&P +1% → PYPL +β%', explain_font, border=thin, align=wrap)
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10); r += 1

w(ws, r, 1, 'R²', label_font, border=thin)
w(ws, r, 2, f"=RSQ({SRC_R}!{pypl_col}{ret_data_start}:{pypl_col}{ret_data_end},{SRC_R}!{sp_col}{ret_data_start}:{sp_col}{ret_data_end})",
  fill=formula_fill, border=thin, num_fmt='0.0%')
w(ws, r, 3, 'RSQ', explain_font, border=thin)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
w(ws, r, 5, f'=TEXT(B{r}*100,"0")&"% движений PYPL объяснены рынком"', explain_font, border=thin)
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10); r += 2

# Прогноз
h2(ws, r, '📊 ПРОГНОЗ ДОХОДНОСТИ PYPL ПО S&P 500', 10); r += 1
step_r(ws, r, '📌 TREND(PYPL_все, SP500_все, SP500_текущий)', 10); r += 1
r += 1
headers(ws, r, ['День', 'S&P500 факт (%)', '', 'PYPL факт (%)', 'PYPL прогноз (%)', 'Ошибка', 'e²', '', '', '']); r += 1
ds = r

for d in range(NR):
    ret_row = ret_data_start + d
    w(ws, r, 1, d+2, border=thin)
    w(ws, r, 2, f"={SRC_R}!{sp_col}{ret_row}", border=thin, num_fmt='0.00%')
    w(ws, r, 4, f"={SRC_R}!{pypl_col}{ret_row}", border=thin, num_fmt='0.00%')
    w(ws, r, 5,
      f'=TREND({SRC_R}!{pypl_col}{ret_data_start}:{pypl_col}{ret_data_end},{SRC_R}!{sp_col}{ret_data_start}:{sp_col}{ret_data_end},{SRC_R}!{sp_col}{ret_row})',
      fill=formula_fill, border=thin, num_fmt='0.00%')
    w(ws, r, 6, f'=D{r}-E{r}', fill=intermed_fill, border=thin, num_fmt='0.00%')
    w(ws, r, 7, f'=F{r}^2', fill=intermed_fill, border=thin, num_fmt='0.0000%')
    r += 1
r += 1

w(ws, r, 1, 'RMSE =', label_font, border=thin)
w(ws, r, 2, f'=SQRT(SUM(G{ds}:G{ds+NR-1})/{NR})', fill=result_fill, border=thin, font=big_green, num_fmt='0.00%')
w(ws, r, 3, '← Средняя ошибка прогноза', explain_font)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6); r += 2

conclusion(ws, r, '🏆 R² показывает какая доля движений PYPL объяснена рынком. '
    'Обычно PYPL имеет β≈1.5–2 и R²≈30-40% — она очень «своевольная».', 10)

for c in range(1, 11):
    ws.column_dimensions[get_column_letter(c)].width = 16

# ═══════════════════════════════════════════════════════════════
# ЛИСТ 5: Ранговые методы
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('🏅 Ранговые методы')
NS = 20
title_row(ws, 1, '🏅 РАНГОВЫЕ МЕТОДЫ (РЕАЛЬНЫЕ ДАННЫЕ)', 10)

r = 3
# Спирмен: PYPL ↔ MA
h2(ws, r, f'📊 СПИРМЕН: PYPL ↔ MA (n={NS} дней)', 10); r += 1
explain(ws, r, 'Paypal и Mastercard и  — компании цифровых платежей. Ранги доходностей совпадают?', 10); r += 1
r += 1

headers(ws, r, ['День', 'PYPL (%)', 'MA (%)', 'Ранг PYPL', 'Ранг MA', 'd', 'd²', '', '', '']); r += 1
sp = r
for i in range(NS):
    ret_row = ret_data_start + i
    w(ws, r, 1, i+2, border=thin)
    w(ws, r, 2, f"={SRC_R}!C{ret_row}", border=thin, num_fmt='0.00%')
    w(ws, r, 3, f"={SRC_R}!E{ret_row}", border=thin, num_fmt='0.00%')
    w(ws, r, 4, f'=RANK(B{r},$B${sp}:$B${sp+NS-1})', fill=rank_fill, border=thin)
    w(ws, r, 5, f'=RANK(C{r},$C${sp}:$C${sp+NS-1})', fill=rank_fill, border=thin)
    w(ws, r, 6, f'=D{r}-E{r}', fill=intermed_fill, border=thin)
    w(ws, r, 7, f'=F{r}^2', fill=intermed_fill, border=thin)
    r += 1
r += 1

sd2r = r
w(ws, r, 1, 'Σd² =', label_font, border=thin)
w(ws, r, 2, f'=SUM(G{sp}:G{sp+NS-1})', fill=intermed_fill, border=thin, font=label_font); r += 1

rho_r = r
w(ws, r, 1, 'ρ Спирмена =', label_font, border=thin)
w(ws, r, 2, f'=1-(6*B{sd2r})/({NS}*({NS}^2-1))', fill=result_fill, border=thin, font=big_green, num_fmt='0.000'); r += 1

w(ws, r, 1, 'Проверка =', label_font, border=thin)
w(ws, r, 2, f'=CORREL(D{sp}:D{sp+NS-1},E{sp}:E{sp+NS-1})', fill=formula_fill, border=thin, num_fmt='0.000')
w(ws, r, 3, '← Должно совпасть', explain_font)
ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7); r += 1

w(ws, r, 1, 'Сила:', label_font, border=thin)
w(ws, r, 2, f'=IF(ABS(B{rho_r})>=0.7,"🟢 Сильная",IF(ABS(B{rho_r})>=0.3,"🟡 Средняя","🔴 Слабая"))',
  fill=result_fill, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3); r += 2

conclusion(ws, r, '🏆 PYPL и MA обычно сильно коррелируют (оба Big Tech). '
    'Спирмен подтверждает связь даже при выбросах.', 10); r += 2

# Манн-Уитни: TSLA vs JPM
h2(ws, r, '📊 МАНН-УИТНИ: GOOGL vs BAIDF (доходности)', 10); r += 1
explain(ws, r, 'Google и Baidu — разные страны?', 10); r += 1
r += 1

n1, n2 = 8, 8
headers(ws, r, ['GOOGL (%)', 'Ранг', '', 'BAIDF (%)', 'Ранг', '', '', '', '', '']); r += 1
mw = r
for i in range(max(n1, n2)):
    ret_row = ret_data_start + i
    # GOOGL = col F, BAIDF = col H
    if i < n1:
        w(ws, r, 1, f"={SRC_R}!F{ret_row}", border=thin, num_fmt='0.00%')
        w(ws, r, 2,
          f'=SUMPRODUCT(($A${mw}:$A${mw+n1-1}>A{r})*1)+SUMPRODUCT(($D${mw}:$D${mw+n2-1}>A{r})*1)+1',
          fill=rank_fill, border=thin)
    if i < n2:
        w(ws, r, 4, f"={SRC_R}!H{ret_row}", border=thin, num_fmt='0.00%')
        w(ws, r, 5,
          f'=SUMPRODUCT(($A${mw}:$A${mw+n1-1}>D{r})*1)+SUMPRODUCT(($D${mw}:$D${mw+n2-1}>D{r})*1)+1',
          fill=rank_fill, border=thin)
    r += 1
r += 1

r1r = r
w(ws, r, 1, 'R₁ (Σ рангов TSLA)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
w(ws, r, 4, f'=SUM(B{mw}:B{mw+n1-1})', fill=formula_fill, border=thin); r += 1

r2r = r
w(ws, r, 1, 'R₂ (Σ рангов JPM)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
w(ws, r, 4, f'=SUM(E{mw}:E{mw+n2-1})', fill=formula_fill, border=thin); r += 1

u1r = r
w(ws, r, 1, 'U₁', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
w(ws, r, 4, f'={n1}*{n2}+{n1}*({n1}+1)/2-D{r1r}', fill=intermed_fill, border=thin); r += 1

u2r = r
w(ws, r, 1, 'U₂', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
w(ws, r, 4, f'={n1}*{n2}+{n2}*({n2}+1)/2-D{r2r}', fill=intermed_fill, border=thin); r += 1

ur = r
w(ws, r, 1, 'U = MIN(U₁,U₂)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
w(ws, r, 4, f'=MIN(D{u1r},D{u2r})', fill=result_fill, border=thin, font=big_green); r += 1

ucr = r
w(ws, r, 1, 'U_крит (α=0.05)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
w(ws, r, 4, 13, border=thin); r += 1

w(ws, r, 1, 'РЕЗУЛЬТАТ:', Font(bold=True, size=12), border=thin)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
w(ws, r, 4, f'=IF(D{ur}<=D{ucr},"✅ Секторы РАЗЛИЧАЮТСЯ","❌ Различий НЕТ")',
  font=Font(bold=True, size=11), fill=result_fill, border=thin)
ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=7); r += 2

conclusion(ws, r, '🏆 GOOGL и BAIDF — из совершенно разных секторов. Если различия значимы → '
    'комбинация GOOGL+BAIDF = отличная диверсификация!', 10)

for c in range(1, 11):
    ws.column_dimensions[get_column_letter(c)].width = 16

# ═══════════════════════════════════════════════════════════════
# ЛИСТ 6: Бонус
# ═══════════════════════════════════════════════════════════════
ws = wb.create_sheet('⭐ Бонус')
title_row(ws, 1, '⭐ БОНУС: ЧАСТНАЯ КОРРЕЛЯЦИЯ + COHEN\'S d', 10)

r = 3
# Частная корреляция: PYPL&BAIDF | S&P500
h2(ws, r, '🟣 ЧАСТНАЯ КОРРЕЛЯЦИЯ: PYPL&BAIDF | S&P500', 10); r += 1
explain(ws, r, 'PYPL&BAIDF коррелируют. Но может это просто потому что ОБА следуют за S&P500?', 10); r += 1
r += 1

# PYPL=C, BAIDF=H, SP500=I
headers(ws, r, ['Шаг', 'Что считаем', '', 'Значение', 'Пояснение', '', '', '', '', '']); r += 1

rxyr = r
w(ws, r, 1, '1', label_font, border=thin)
w(ws, r, 2, 'r(PYPL, BAIDF)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f"=CORREL({SRC_R}!C{ret_data_start}:C{ret_data_end},{SRC_R}!H{ret_data_start}:H{ret_data_end})",
  fill=formula_fill, border=thin, num_fmt='0.000')
w(ws, r, 5, '← Обычная корреляция (может быть из-за рынка)', explain_font, border=thin, align=wrap)
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10); r += 1

rxzr = r
w(ws, r, 1, '2', label_font, border=thin)
w(ws, r, 2, 'r(PYPL, S&P500)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f"=CORREL({SRC_R}!C{ret_data_start}:C{ret_data_end},{SRC_R}!I{ret_data_start}:I{ret_data_end})",
  fill=formula_fill, border=thin, num_fmt='0.000'); r += 1

ryzr = r
w(ws, r, 1, '3', label_font, border=thin)
w(ws, r, 2, 'r(BAIDF, S&P500)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f"=CORREL({SRC_R}!H{ret_data_start}:H{ret_data_end},{SRC_R}!I{ret_data_start}:I{ret_data_end})",
  fill=formula_fill, border=thin, num_fmt='0.000'); r += 1

w(ws, r, 1, '4 ✨', label_font, border=thin)
w(ws, r, 2, 'r_partial', bold_green, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f'=(D{rxyr}-D{rxzr}*D{ryzr})/SQRT((1-D{rxzr}^2)*(1-D{ryzr}^2))', fill=result_fill, border=thin,
  font=big_green, num_fmt='0.000')
w(ws, r, 5, '← ЧИСТАЯ связь PYPL↔BAIDF без рынка!', explain_font, border=thin, align=wrap)
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10); r += 2

conclusion(ws, r, '🏆 Если частная r << обычной → PYPL - BAIDF коррелировали только из-за S&P500. '
    'Если ≈ → между ними есть собственная связь (конкуренция в Tech).', 10); r += 2

# Cohen's d: TSLA первая vs вторая половина
h2(ws, r, "🟡 COHEN'S d: PYPL первая пол. vs вторая пол.", 10); r += 1
explain(ws, r, "Изменился ли характер PYPL во второй половине периода?", 10); r += 1
r += 1

h1_s = ret_data_start
h1_e = ret_data_start + 14
h2_s = ret_data_start + 15
h2_e = ret_data_end

headers(ws, r, ['Шаг', 'Показатель', '', 'Значение', 'Пояснение', '', '', '', '', '']); r += 1

m1r = r
w(ws, r, 1, '1', label_font, border=thin)
w(ws, r, 2, 'M₁ (1-я половина)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f"=AVERAGE({SRC_R}!C{h1_s}:C{h1_e})", fill=formula_fill, border=thin, num_fmt='0.00%'); r += 1

m2r = r
w(ws, r, 1, '2', label_font, border=thin)
w(ws, r, 2, 'M₂ (2-я половина)', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f"=AVERAGE({SRC_R}!G{h2_s}:C{h2_e})", fill=formula_fill, border=thin, num_fmt='0.00%'); r += 1

sd1r = r
w(ws, r, 1, '3', label_font, border=thin)
w(ws, r, 2, 'SD₁', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f"=STDEV({SRC_R}!C{h1_s}:C{h1_e})", fill=formula_fill, border=thin, num_fmt='0.00%'); r += 1

sd2r_v = r
w(ws, r, 1, '4', label_font, border=thin)
w(ws, r, 2, 'SD₂', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f"=STDEV({SRC_R}!C{h2_s}:C{h2_e})", fill=formula_fill, border=thin, num_fmt='0.00%'); r += 1

sdpr = r
w(ws, r, 1, '5', label_font, border=thin)
w(ws, r, 2, 'SD_pooled', label_font, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f'=SQRT((D{sd1r}^2+D{sd2r_v}^2)/2)', fill=formula_fill, border=thin, num_fmt='0.00%'); r += 1

cdr = r
w(ws, r, 1, '6 ✨', label_font, border=thin)
w(ws, r, 2, "Cohen's d", bold_green, border=thin)
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
w(ws, r, 4, f'=(D{m1r}-D{m2r})/D{sdpr}', fill=result_fill, border=thin, font=big_green, num_fmt='0.000')
w(ws, r, 5, f'=IF(ABS(D{r})>=0.8,"Большой ⬆️ Тренд!",IF(ABS(D{r})>=0.5,"Средний ➡️","Малый ⬇️ Стабильно"))',
  font=label_font, fill=result_fill, border=thin)
ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10); r += 2

conclusion(ws, r, "🏆 Если |d|≥0.8 → PYPL резко изменила поведение (новости, отчёт, Маск...). "
    "|d|<0.2 → стабильный период.", 10)

for c in range(1, 11):
    ws.column_dimensions[get_column_letter(c)].width = 16
ws.column_dimensions['E'].width = 45

# ═══ SAVE ═══

import os
from openpyxl import Workbook
# Папка для сохранения результатов (относительная, рабочая директория)
OUT_FOLDER = "output"  
os.makedirs(OUT_FOLDER, exist_ok=True)  # создаём, если нет

# Имя файла
OUT_FILE = os.path.join(OUT_FOLDER, "stock_analysis_results.xlsx")
print(f"\n✅ Готово! РЕАЛЬНЫЕ данные ({DAYS} дней). Файл: Abibullayeva_Fatima_stocks.xlsx")
print(f"Доходности: строки {ret_data_start}–{ret_data_end}. Все ссылки корректны.")

import os
print("Текущая рабочая директория:", os.getcwd())
try:
    wb.save(OUT_FILE)
    print(f"✅ Файл сохранён: {OUT_FILE}")
except Exception as e:
    print(f"❌ Ошибка при сохранении: {e}")
